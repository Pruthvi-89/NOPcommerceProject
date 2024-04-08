import openpyxl

def getrowcount(file,sheetname):
    book=openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return (sheet.max_row)



def readdata(file,sheetname,rowno,colno):
    book=openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.cell(row=rowno,column= colno).value


def writedata(file,sheetname,rowno,colno,data):
    """

    :rtype: object
    """
    book=openpyxl.load_workbook(file)
    sheet = book[sheetname]
    sheet.cell(row=rowno,column=colno).value = data
    book.save(file)